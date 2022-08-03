from bs4 import BeautifulSoup
import requests
import random
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from time import sleep
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO
import re

session = requests.session()

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) '
                  'Gecko/20100101 Firefox/88.0',
    'Accept': '*/*',
    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
    'Accept-Encoding': 'gzip, deflate, br',
    # 'Content-Type': 'application/json',
    'X-Requested-With': 'XMLHttpRequest',
    'x-date-format': 'iso',
    #'Content-Length': '220',
    'Origin': 'https://kad.arbitr.ru',
    'Connection': 'keep-alive',
    'Referer': 'https://kad.arbitr.ru/',
    'Cookie':
    'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=3cad84fc-05e8-4234-9afd-6f2fcc85dad9:5EjdVHanx18mgJ+dYKJadw==; _ga=GA1.2.133637717.1616925617; _gid=GA1.2.2138084442.1616925617; _ym_uid=1616925617387222795; _ym_d=1616925617; tmr_reqNum=17; tmr_lvid=e4b171a78745d501bf31deec258bfbfd; tmr_lvidTS=1616925616977; _fbp=fb.1.1616925617005.1859496784; _ym_isad=2; tmr_detect=0%7C1616928315811; ASP.NET_SessionId=bwalesrby4fceg3yzsezukti; _gat=1; _gat_FrontEndTracker=1; _dc_gtm_UA-157906562-1=1; _ym_visorc=b; pr_fp=70e21f14878767bdf2684e7eb6a5307017b708c45b3c2a0080889f792151e0d4; wasm=6b99641fa4d6b22f646c77d2ed223d9a; rcid=9d9d4a78-f271-4ca3-b201-e0fabcf71a49'
}

headersget = {
'Host': 'kad.arbitr.ru',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) '
              'Gecko/20100101 Firefox/88.0',
'Accept': 'text/javascript, application/javascript, */*',
'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
'Accept-Encoding': 'gzip, deflate, br',
'X-Requested-With': 'XMLHttpRequest',
'Connection': 'keep-alive',
'Referer': 'https://kad.arbitr.ru/',
'Cookie': 'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=3cad84fc-05e8-4234-9afd-6f2fcc85dad9:5EjdVHanx18mgJ+dYKJadw==; _ga=GA1.2.133637717.1616925617; _ym_uid=1616925617387222795; _ym_d=1616925617; tmr_reqNum=68; tmr_lvid=e4b171a78745d501bf31deec258bfbfd; tmr_lvidTS=1616925616977; _fbp=fb.1.1616925617005.1859496784; ASP.NET_SessionId=bwalesrby4fceg3yzsezukti; pr_fp=70e21f14878767bdf2684e7eb6a5307017b708c45b3c2a0080889f792151e0d4; rcid=9d9d4a78-f271-4ca3-b201-e0fabcf71a49; .ASPXAUTH=E0C43EDF7307DB05145231143A672FF39CD9E67CDF3FCC784A2570D0DED2433A31C4EDAE8402FB21C1BE3A4FE7ACE8D943EF3E62350877A8BECEC68CA7106DF646CA004F73A3A33CE297C1455F3836CDA37C3EF569A15A097CCC2D518BD73689FC9D8A91; _gid=GA1.2.1689069161.1617723817; _ym_isad=2; tmr_detect=0%7C1617738655382; KadLVCards=%d0%9041-30373%2f2018~%d0%9041-54471%2f2019~%d0%9041-67962%2f2019~%d0%9041-73264%2f2020~%d0%9004-2427%2f2021; _dc_gtm_UA-157906562-1=1; _gat=1; _gat_FrontEndTracker=1'
}

headersget_2 = {
'Host': 'kad.arbitr.ru',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) '
              'Gecko/20100101 Firefox/88.0',
'Accept': '*/*',
'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
'Accept-Encoding': 'gzip, deflate, br',
'Referer': 'https://kad.arbitr.ru/',
'Connection': 'keep-alive',
'Cookie': 'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=3cad84fc-05e8-4234-9afd-6f2fcc85dad9:5EjdVHanx18mgJ+dYKJadw==; _ga=GA1.2.133637717.1616925617; _ym_uid=1616925617387222795; _ym_d=1616925617; tmr_reqNum=68; tmr_lvid=e4b171a78745d501bf31deec258bfbfd; tmr_lvidTS=1616925616977; _fbp=fb.1.1616925617005.1859496784; ASP.NET_SessionId=bwalesrby4fceg3yzsezukti; pr_fp=70e21f14878767bdf2684e7eb6a5307017b708c45b3c2a0080889f792151e0d4; rcid=9d9d4a78-f271-4ca3-b201-e0fabcf71a49; .ASPXAUTH=E0C43EDF7307DB05145231143A672FF39CD9E67CDF3FCC784A2570D0DED2433A31C4EDAE8402FB21C1BE3A4FE7ACE8D943EF3E62350877A8BECEC68CA7106DF646CA004F73A3A33CE297C1455F3836CDA37C3EF569A15A097CCC2D518BD73689FC9D8A91; _gid=GA1.2.1689069161.1617723817; _ym_isad=2; tmr_detect=0%7C1617738655382; KadLVCards=%d0%9041-30373%2f2018~%d0%9041-54471%2f2019~%d0%9041-67962%2f2019~%d0%9041-73264%2f2020~%d0%9004-2427%2f2021; _dc_gtm_UA-157906562-1=1; _gat=1; _gat_FrontEndTracker=1'
}

headerspost = {
'Host': 'kad.arbitr.ru',
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) '
              'Gecko/20100101 Firefox/88.0',
'Accept': '*/*',
'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
'Accept-Encoding': 'gzip, deflate, br',
'Content-Type': 'application/json',
'X-Requested-With': 'XMLHttpRequest',
'x-date-format': 'iso',
'Content-Length': '215',
'Origin': 'https://kad.arbitr.ru',
'Connection': 'keep-alive',
'Referer': 'https://kad.arbitr.ru/',
'Cookie': 'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=3cad84fc-05e8-4234-9afd-6f2fcc85dad9:5EjdVHanx18mgJ+dYKJadw==; _ga=GA1.2.133637717.1616925617; _ym_uid=1616925617387222795; _ym_d=1616925617; tmr_reqNum=68; tmr_lvid=e4b171a78745d501bf31deec258bfbfd; tmr_lvidTS=1616925616977; _fbp=fb.1.1616925617005.1859496784; ASP.NET_SessionId=bwalesrby4fceg3yzsezukti; pr_fp=70e21f14878767bdf2684e7eb6a5307017b708c45b3c2a0080889f792151e0d4; rcid=9d9d4a78-f271-4ca3-b201-e0fabcf71a49; .ASPXAUTH=E0C43EDF7307DB05145231143A672FF39CD9E67CDF3FCC784A2570D0DED2433A31C4EDAE8402FB21C1BE3A4FE7ACE8D943EF3E62350877A8BECEC68CA7106DF646CA004F73A3A33CE297C1455F3836CDA37C3EF569A15A097CCC2D518BD73689FC9D8A91; _gid=GA1.2.1689069161.1617723817; _ym_isad=2; tmr_detect=0%7C1617738655382; KadLVCards=%d0%9041-30373%2f2018~%d0%9041-54471%2f2019~%d0%9041-67962%2f2019~%d0%9041-73264%2f2020~%d0%9004-2427%2f2021; _dc_gtm_UA-157906562-1=1; _gat=1; _gat_FrontEndTracker=1; wasm=b9db58eb07444548543ae605d70acd8b'
}

# FIO = ''  # ФИО искомого человека
# -1 - Любой
# 0 - Истец
# 1 - Ответчик
# 2 - Третье лицо
# 3 - Иное лицо
role = 1    # Роль искомого лица в процессе

req = session.get('https://kad.arbitr.ru/', headers=headers)

get1 = session.get('https://kad.arbitr.ru/Wasm/api/v1/wasm.js?_=1617738668336', headers=headersget)
get2 = session.get('https://kad.arbitr.ru/Wasm/api/v1/wasm_bg.wasm?_'
                   '=1617738668355', headers=headersget_2)

def PDF_res(path):

    def convert_pdf_to_txt(path_1):
        path = path_1
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = open(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0
        caching = True
        pagenos = set()

        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                      password=password, caching=caching,
                                      check_extractable=True):
            interpreter.process_page(page)

        text = retstr.getvalue()

        fp.close()
        device.close()
        retstr.close()
        return text

    pattern = r'(?:([Р|р]\s*[Е|е]\s*[Ш|ш]\s*[И|и]\s*[Л|л]\s*\:?)|([П|п]\s*[О|о]\s*[С|с]\s*[Т|т]\s*[А|а]\s*[Н|н]\s*[О|о]\s*[В|в]\s*[И|и]\s*[Л|л]\s*\:?)|([О|о]\s*[П|п]\s*[Р|р]\s*[Е|е]\s*[Д|д]\s*[Е|е]\s*[Л|л]\s*[И|и]\s*[Л|л]\s*\:?))(.*\s*)*'

    def get_res(path2):
        global pattern
        text_pdf = (convert_pdf_to_txt(path2))
        pdf_words = text_pdf.split()
        text = ''
        for i in range(len(pdf_words)):
            if i != 0 and i % 10 == 0:
                text = text + ' ' + pdf_words[i] + '\n'
            else:
                text = text + ' ' + pdf_words[i]

        try:
            text = re.search(pattern, text)
            text = text.group(0)
        except Exception:
            text = "В документе совпадений по шаблону не найдено"
        return text

    return get_res(path)


def get_arb_data(FIO, role):

    global headerspost
    page = 1
    result_data = ''
    while True:
        params = {"Page": page, "Count": 25, "Courts": [], "DateFrom": "null",
                  "DateTo": 'null',
                  "Sides": [{"Name": FIO, "Type": role,
                             "ExactMatch": 'true'}],
                  "Judges": [], "CaseNumbers": [], "WithVKSInstances": 'false'}

        try:
            post1 = session.post('https://kad.arbitr.ru/Kad/SearchInstances', headers=headerspost, json=params)
            t = random.random() + random.randrange(3, 5)
            sleep(t)

        except Exception:
            print('Ошибка данных')
            break
        page += 1
        if len(post1.text) <= 242:
            break
        result_data = result_data + post1.text + '\n'


    soup = BeautifulSoup(result_data, 'lxml')
    cells = soup.find_all('div', class_='b-container')
    count = 0
    my_dict = {}
    numb = 1
    for cell in cells:
        my_list = []
        count += 1
        cell_soup = cell
        # 1-я ячейка
        if count == 1:
            try:
                cell_A = cell_soup.find('span').text.strip()
                if len(cell_A) == 0:
                    cell_A = 'Данные отсутствуют'
            except Exception:
                cell_A = 'Нет данных'
            try:
                cell_B = cell_soup.find('a').text.strip()
                if len(cell_B) == 0:
                    cell_B = 'Данные отсутствуют'
            except Exception:
                cell_B = 'Нет данных'
            try:
                cell_B_href = cell_soup.find('a').get('href')
                t = random.random() + random.randrange(2, 4)
                sleep(t)
                if len(cell_B_href) == 0:
                    cell_B_href = 'Данные отсутствуют'
            except Exception:
                cell_B_href = 'Нет данных'
            B_name = cell_B
            cell_B = f'=HYPERLINK("{cell_B_href}","{cell_B}")'
            try:
                cell_G = PDF_res(cell_B_href)
                if len(cell_G) == 0:
                    cell_G = 'Ошибка обработки данных'
            except Exception:
                cell_G = 'Error'
        # 2-я ячейка
        elif count == 2:
            try:
                cell_C = cell_soup.find(class_='judge').text.strip()
                if len(cell_C) == 0:
                    cell_C = 'Данные отсутствуют'
            except Exception:
                cell_C = 'Нет данных'
            try:
                cell_D = cell_soup.find('div', class_=False).text.strip()
                if len(cell_D) == 0:
                    cell_D = 'Данные отсутствуют'
            except Exception:
                cell_D = 'Нет данных'
        # 3-я ячейка
        elif count == 3:
            try:
                cell_E_all = cell_soup.find_all(class_='js-rollover b-newRollover')
                cell_E = ''
                for elem in cell_E_all:
                    elem = elem.text.strip()
                    try:
                        elem = list(elem.split('\n'))
                        elem = elem[0].strip()
                        cell_E = cell_E + elem + '\n'
                    except Exception as ex:
                        cell_E = ex
                if len(cell_E) == 0:
                    cell_E = 'Данные отсутствуют'
            except Exception as ex:
                cell_E = ex
        # 4-я ячейка
        elif count == 4:
            try:
                cell_F_all = cell_soup.find_all(class_='js-rollover b-newRollover')
                cell_F = ''
                for elem2 in cell_F_all:
                    elem2 = elem2.text.strip()
                    try:
                        elem2 = list(elem2.split('\n'))
                        elem2 = elem2[0].strip() + '\n' + elem2[-1].strip()
                        cell_F = cell_F + elem2 + '\n'
                    except Exception as ex:
                        cell_F = ex
                if len(cell_F) == 0:
                    cell_F = 'Данные отсутствуют'
            except Exception as ex:
                cell_F = ex

            my_list.append([cell_A, cell_B, cell_C, cell_D, cell_E, cell_F,
                            cell_G, cell_B_href, B_name])
            my_dict[numb] = my_list
            numb += 1
            count = 0
        else:
            my_dict['1'] = 'ОШИБКА'

    return my_dict


def get_excel_table(FIO):
    global get_arb_data
    global role
    res_dict = get_arb_data(FIO, role)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'Дата'
    sheet['A1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['B1'] = 'Дело'
    sheet['B1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['C1'] = 'Судья'
    sheet['C1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['D1'] = 'Текущая инстанция'
    sheet['D1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['E1'] = 'Истец'
    sheet['E1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['F1'] = 'Ответчик'
    sheet['F1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    sheet['G1'] = 'Решение из PDF'
    sheet['G1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
    book_name = FIO + '.xlsx'

    wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    wrap_alignment2 = Alignment(wrap_text=True, horizontal='left', vertical='center')
    sheet.cell(row=1, column=1).alignment = wrap_alignment
    sheet.cell(row=1, column=2).alignment = wrap_alignment
    sheet.cell(row=1, column=3).alignment = wrap_alignment
    sheet.cell(row=1, column=4).alignment = wrap_alignment
    sheet.cell(row=1, column=5).alignment = wrap_alignment
    sheet.cell(row=1, column=6).alignment = wrap_alignment
    sheet.cell(row=1, column=7).alignment = wrap_alignment

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 17
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 70
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 70

    sheet_names_list = []
    href_list = []
    for key, value in res_dict.items():
        row_number = key + 1

        sheet.cell(row=row_number, column=1).value = value[0][0]
        sheet.cell(row=row_number, column=1).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=2).value = value[0][1]
        sheet.cell(row=row_number, column=2).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=2).font = Font(color='0049FF')
        sheet.cell(row=row_number, column=3).value = value[0][2]
        sheet.cell(row=row_number, column=3).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=4).value = value[0][3]
        sheet.cell(row=row_number, column=4).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=5).value = value[0][4]
        sheet.cell(row=row_number, column=5).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=6).value = value[0][5]
        sheet.cell(row=row_number, column=6).alignment = wrap_alignment2
        sheet.cell(row=row_number, column=7).value = value[0][6]
        sheet.cell(row=row_number, column=7).alignment = wrap_alignment2

        sheet_name = value[0][8]
        sheet_name = sheet_name.replace('/', '_')
        sheet_names_list.append(sheet_name)
        href_list.append(value[0][7])
    c = sheet['A2']
    sheet.freeze_panes = c


    def get_case(case_href, sheet_name):

        session2 = requests.session()
        nonlocal wrap_alignment
        nonlocal book
        nonlocal wrap_alignment2

        headersget_1 = {
            "Host": "kad.arbitr.ru",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0",
            "Accept": "*/*",
            "Accept-Language": "ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Referer": case_href,
            "Cookie": 'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=ea7784d7-e43d-41d8-9e23-c09e37e18c8f:FCoUTtltMpy2i3a/kdD4Mw==; _ga=GA1.2.1625590012.1617866332; _ym_uid=1617866332852707217; _ym_d=1617866332; tmr_reqNum=28; tmr_lvid=6b07a7f6856c3fe127d01e06b53864cc; tmr_lvidTS=1617866331895; _fbp=fb.1.1617866331948.1103016667; ASP.NET_SessionId=t3oyrmdoahzc3zub5dddh5mu; _gid=GA1.2.1173569852.1617977500; _ym_isad=2; pr_fp=74d4c2fde42cef727180ca1b64fdb092b94895598722b90248a56d88b26485f3; tmr_detect=0%7C1617979466666; KadLVCards=%d0%9041-31726%2f2020; wasm=097b543baed650b9f00d2ceb83f32654; rcid=224bdc40-7379-44e3-8592-0fc8580343f7; _gat=1; _gat_FrontEndTracker=1; _dc_gtm_UA-157906562-1=1',
            "Cache-Control": "max-age=0"
        }

        get_1 = session2.get('https://kad.arbitr.ru/Content/Static/Js/Common'
                             '/layout.202103242043.js',
                            headers=headersget_1)
        t = random.random() + random.randrange(3, 5)
        sleep(t)


        main_headers = {
            'Host': 'kad.arbitr.ru',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) '
                          'Gecko/20100101 Firefox/88.0',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,'
                      'image/webp,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate, br',
            'Referer': 'https://kad.arbitr.ru/',
            'Connection': 'keep-alive',
            'Cookie': 'notShowTooltip=yes; '
                      'Notification_All=b4065a6530a34322bb59dae9c84e55ed_1618261200000_shown; CUID=3cad84fc-05e8-4234-9afd-6f2fcc85dad9:5EjdVHanx18mgJ+dYKJadw==; _ga=GA1.2.133637717.1616925617; _ym_uid=1616925617387222795; _ym_d=1616925617; tmr_reqNum=127; tmr_lvid=e4b171a78745d501bf31deec258bfbfd; tmr_lvidTS=1616925616977; _fbp=fb.1.1616925617005.1859496784; _gid=GA1.2.1689069161.1617723817; KadLVCards=%d0%9041-31726%2f2020~%d0%9008-134%2f2016~%d0%9040-79531%2f2015~%d0%9014-18542%2f2009~%d0%9035-5991%2f2006; .ASPXAUTH=E0C43EDF7307DB05145231143A672FF39CD9E67CDF3FCC784A2570D0DED2433A31C4EDAE8402FB21C1BE3A4FE7ACE8D943EF3E62350877A8BECEC68CA7106DF646CA004F73A3A33CE297C1455F3836CDA37C3EF569A15A097CCC2D518BD73689FC9D8A91; ASP.NET_SessionId=phsmnu4wpxc3ni2o1iwqr0j5; rcid=90e68aa6-a8af-43e0-af4c-d7cbd79a8eba; pr_fp=70e21f14878767bdf2684e7eb6a5307017b708c45b3c2a0080889f792151e0d4; _ym_isad=2; tmr_detect=0%7C1618059046806; wasm=cf307a73bed3d35e5c80b71b4f25ca1e; _gat=1; _gat_FrontEndTracker=1; _dc_gtm_UA-157906562-1=1',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0'
        }

        req = session2.get(case_href, headers=main_headers)
        t = random.random() + random.randrange(3, 5)
        sleep(t)

        # with open('res.html', 'w', encoding='utf-8') as fout:
        #     fout.write(req.text)

        case_soup = BeautifulSoup(req.text, 'lxml')
        instances = case_soup.find_all('div', class_='l-col')
        instances_list = []  # Список типов инстанций
        for instance in instances:
            instance = instance.text.strip()
            instance_l = list(instance.split('\n'))
            instance = instance_l[0].strip() + ' ' + instance_l[-1].strip()
            instances_list.append(instance)

        instance_numbers = case_soup.find_all('strong', class_="b-case-instance-number")
        instance_numbers_list = []  # Список номеров инстанций
        for ins_numb in instance_numbers:
            instance_numbers_list.append(ins_numb.text.strip())

        instantion_names = case_soup.find_all('span', class_='instantion-name')
        instantion_names_list = []  # Список наименований судов
        instantion_names_href_list = []  # Список ссылок на суды
        for inst_name in instantion_names:
            instantion_names_list.append(inst_name.text.strip())
            instantion_names_href_list.append(inst_name.find('a').get('href'))

        case_results = case_soup.find_all('h2', class_='b-case-result')
        case_result_name_list = []  # Название документа решения суда
        case_result_href_list = []  # Ссылка на документ решения суда
        for case_res in case_results:
            case_result_name_list.append(case_res.text.strip())
            try:
                case_result_href_list.append(case_res.find('a').get('href'))
            except Exception:
                case_result_href_list.append('Документ заседания отсутствует')

        sheet_name = f'Дело {sheet_name}'
        case_sheet = book.create_sheet(title=sheet_name)
        case_sheet = book[sheet_name]
        case_sheet['A1'] = 'Дата и вид инстанции'
        case_sheet['A1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
        case_sheet['B1'] = 'Номер инстанции'
        case_sheet['B1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
        case_sheet['C1'] = 'Наименование суда'
        case_sheet['C1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
        case_sheet['D1'] = 'Ссылка на документ'
        case_sheet['D1'].fill = PatternFill(start_color="86FF8C", end_color="86FF8C", fill_type="solid")
        case_sheet.cell(row=1, column=1).alignment = wrap_alignment
        case_sheet.cell(row=1, column=2).alignment = wrap_alignment
        case_sheet.cell(row=1, column=3).alignment = wrap_alignment
        case_sheet.cell(row=1, column=4).alignment = wrap_alignment
            # instances_list = []  # Список типов инстанций
            # instance_numbers_list = []  # Список номеров инстанций
            # instantion_names_list = []  # Список наименований судов
            # instantion_names_href_list = []  # Список ссылок на суды
            # case_result_name_list = []  # Название документа решения суда
            # case_result_href_list = []  # Ссылка на документ решения суда
        count_row = 1
        for i in range(len(instances_list)):
            count_row += 1
            case_sheet.cell(row=count_row, column=1).value = instances_list[i]
            case_sheet.cell(row=count_row, column=1).alignment = wrap_alignment2
            case_sheet.cell(row=count_row, column=2).value = instance_numbers_list[i]
            case_sheet.cell(row=count_row, column=2).alignment = wrap_alignment2
            case_sheet.cell(row=count_row, column=3).value = f'=HYPERLINK("{instantion_names_href_list[i]}","{instantion_names_list[i]}")'
            case_sheet.cell(row=count_row, column=3).alignment = wrap_alignment2
            case_sheet.cell(row=count_row, column=3).font = Font(color='0049FF')
            try:
                if case_result_href_list[i] == 'Документ заседания ' \
                                               'отсутствует':
                    case_sheet.cell(row=count_row,
                                    column=4).value = 'Документ ' \
                                                      'заседания' \
                                                      'отсутствует'
                case_sheet.cell(row=count_row, column=4).value = f'=HYPERLINK("{case_result_href_list[i]}","{case_result_name_list[i]}")'
            except Exception:
                case_sheet.cell(row=count_row, column=4).value = 'Документ ' \
                                                                 'заседания' \
                                                                 'отсутствует'
            case_sheet.cell(row=count_row, column=4).alignment = wrap_alignment2
            case_sheet.cell(row=count_row, column=4).font = Font(color='0049FF')
            case_sheet.column_dimensions['A'].width = 25
            case_sheet.column_dimensions['B'].width = 17
            case_sheet.column_dimensions['C'].width = 35
            case_sheet.column_dimensions['D'].width = 80
        c = case_sheet['A2']
        sheet.freeze_panes = c
        print(f'Лист {sheet_name} - готов')

    k = 0
    for case_href in href_list:
        get_case(case_href, sheet_names_list[k])
        k += 1

    book.save(book_name)
    print(book_name[:-5], ' - Готов')
    book.close()


names_list = input('Введите имя файла со списком людей \n Пример: Список.txt  -  : ')
with open(names_list, 'r', encoding='utf-8') as names:
    lines = names.readlines()
    for line in lines:
        line = line.strip()
        if __name__ == "__main__":
            get_excel_table(line)
